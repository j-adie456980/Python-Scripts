from copy import copy
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, Fill

wb = load_workbook('data.xlsx')
source = wb.active
source.title = "Source"

targetWB = Workbook()
target = targetWB.active

red = source['G5'].fill.start_color

for row in source.iter_rows(max_row=11, max_col=9):
    for cell in row:
        if (cell.value == None):
            temp = chr(cell.column+64) + str(cell.row)
            #print(temp)
            source[temp] = "N/A"


def insertRow(n, i):  #produce the list of items in the particular row
        for row in source.iter_rows(min_row=n, max_row=n, max_col=9):
            target.append((cell.value for cell in row))
            for cell in row:
                temp = chr(cell.column+64) + str(i)
                target[temp].fill = copy(cell.fill)
            
i = 1
for row in source.iter_rows(max_row=11, max_col=9):
    for cell in row:
        if (cell.fill.start_color == red):
            insertRow(cell.row, i)
            i += 1
            break


target.insert_rows(1)
for row in source.iter_rows(max_row=1, max_col=9):
    for cell in row:
        temp = chr(cell.column+64) + str(cell.row)
        target[temp] = copy(cell.value)
targetWB.save('Filtered.xlsx')
